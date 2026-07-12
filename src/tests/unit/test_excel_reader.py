"""The Excel pre-reader against a deliberately ugly synthetic workbook.

The frozen corpus files are dirty in *some* ways (numeric IDs, serial
dates); the generic capability — merged headers, decimal-comma text,
blank rows — is exercised here, domain-free.
"""

import duckdb
import openpyxl

from before_we_ai.sources.excel import (
    RULE_BLANK_ROWS,
    RULE_MERGED_HEADER,
    read_workbook,
    sheet_to_parquet,
)


def ugly_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kunden Liste"
    # Merged parent header over two sub-headers, plus a vertical merge.
    ws["A1"] = "Kunde"
    ws.merge_cells("A1:B1")
    ws["C1"] = "Umsatz 2025"
    ws.merge_cells("C1:C2")
    ws["A2"] = "Nr"
    ws["B2"] = "Ort"
    ws.append([1042, "Köln", "1.234,56"])
    ws.append([None, None, None])  # blank row in the middle
    ws.append([7, " Bonn ", 99.0])
    wb.save(path)
    return path


def test_reader_survives_the_mess(tmp_path):
    sheets = read_workbook(ugly_workbook(tmp_path / "ugly.xlsx"))
    assert len(sheets) == 1
    sheet = sheets[0]
    assert sheet.sheet == "kunden_liste"
    assert sheet.columns == ["kunde_nr", "kunde_ort", "umsatz_2025"]
    assert sheet.rows == [
        ["1042", "Köln", "1234.56"],
        ["7", "Bonn", "99"],
    ]
    rules = {d["rule"] for d in sheet.decisions}
    assert RULE_MERGED_HEADER in rules
    assert RULE_BLANK_ROWS in rules
    assert "decimal_comma_to_dot" in rules
    assert "numeric_to_text" in rules


def test_simple_sheet_single_header_row(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name"])
    ws.append(["0001042", "Muster GmbH"])  # T1: text keeps its zeros
    path = tmp_path / "plain.xlsx"
    wb.save(path)

    sheet = read_workbook(path)[0]
    assert sheet.columns == ["id", "name"]
    assert sheet.rows == [["0001042", "Muster GmbH"]]
    assert all(d["rule"] not in ("numeric_to_text",) for d in sheet.decisions)


def test_parquet_round_trip_is_all_text(tmp_path):
    sheets = read_workbook(ugly_workbook(tmp_path / "ugly.xlsx"))
    con = duckdb.connect()
    out = sheet_to_parquet(con, sheets[0], tmp_path / "cache" / "ugly.parquet")
    types = {r[0]: r[1] for r in con.execute(f"DESCRIBE SELECT * FROM '{out}'").fetchall()}
    assert set(types.values()) == {"VARCHAR"}
    rows = con.execute(f"SELECT * FROM '{out}' ORDER BY kunde_nr").fetchall()
    assert rows == [("1042", "Köln", "1234.56"), ("7", "Bonn", "99")]
