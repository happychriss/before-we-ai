"""The Excel pre-reader (T9): ugly workbooks never go straight to SQL.

openpyxl reads the workbook, merged header cells are resolved, blank
rows are skipped, and every cell is emitted in canonical *text* form
("Typen als Text bewahren") — typed interpretation is the profiler's
job, not the reader's. The result lands as Parquet in ``cache/`` and the
list of applied normalization decisions goes back to the caller for
declaration logging.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from before_we_ai.sources.canonical import canonicalize

RULE_MERGED_HEADER = "merged_header_resolved"
RULE_BLANK_ROWS = "blank_rows_skipped"


@dataclass
class SheetData:
    sheet: str
    columns: list[str]
    rows: list[list[str | None]]
    decisions: list[dict] = field(default_factory=list)


def _sanitize(name: str) -> str:
    slug = re.sub(r"[^0-9a-zA-Z]+", "_", name.strip()).strip("_").lower()
    return slug or "column"


def _dedupe(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for name in names:
        seen[name] = seen.get(name, 0) + 1
        out.append(name if seen[name] == 1 else f"{name}_{seen[name]}")
    return out


def _header_merges(ws, header_row: int) -> dict[int, str]:
    """Map column index -> top-left value for horizontal merges in the header row."""
    filled: dict[int, str] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= header_row <= rng.max_row and rng.max_col > rng.min_col:
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            if top_left is None:
                continue
            for col in range(rng.min_col, rng.max_col + 1):
                filled[col] = str(top_left)
    return filled


def read_workbook(path: str | Path) -> list[SheetData]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        header_idx = next(
            (i for i, row in enumerate(grid) if any(v is not None for v in row)), None
        )
        if header_idx is None:
            continue
        decisions: list[dict] = []
        header = grid[header_idx]
        merges = _header_merges(ws, header_idx + 1)  # openpyxl rows are 1-based

        if merges and len(grid) > header_idx + 1:
            # Classic merged-header layout: a parent cell spans several
            # columns, the sub-headers sit in the following row.
            sub = grid[header_idx + 1]
            names = []
            for i in range(len(header)):
                parent = merges.get(i + 1, header[i])
                child = sub[i] if i < len(sub) else None
                if parent is not None and child is not None:
                    names.append(f"{parent}_{child}")
                else:
                    names.append(str(parent if parent is not None else child or f"col{i + 1}"))
            data_start = header_idx + 2
            decisions.append({
                "rule": RULE_MERGED_HEADER,
                "column": "*",
                "example": {"before": str(header), "after": str(names)},
            })
        else:
            names = [str(v) if v is not None else f"col{i + 1}" for i, v in enumerate(header)]
            data_start = header_idx + 1

        columns = _dedupe([_sanitize(n) for n in names])

        rows: list[list[str | None]] = []
        blank = 0
        rule_seen: dict[tuple[str, str], dict] = {}
        for raw in grid[data_start:]:
            if all(v is None for v in raw):
                blank += 1
                continue
            row = []
            for i, col in enumerate(columns):
                value = raw[i] if i < len(raw) else None
                canon, rule = canonicalize(value)
                if rule and (col, rule) not in rule_seen:
                    rule_seen[(col, rule)] = {
                        "rule": rule,
                        "column": col,
                        "example": {"before": repr(value), "after": canon},
                    }
                row.append(canon)
            rows.append(row)
        if blank:
            decisions.append({
                "rule": RULE_BLANK_ROWS,
                "column": "*",
                "example": {"before": f"{blank} blank rows", "after": "dropped"},
            })
        decisions.extend(rule_seen.values())
        sheets.append(SheetData(_sanitize(ws.title), columns, rows, decisions))
    return sheets


def sheet_to_parquet(con, sheet: SheetData, out_path: str | Path) -> Path:
    """Write one sheet as an all-VARCHAR Parquet file via DuckDB."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    cols = ", ".join(f'"{c}" VARCHAR' for c in sheet.columns)
    con.execute(f"CREATE OR REPLACE TEMP TABLE _sheet_buffer ({cols})")
    if sheet.rows:
        placeholders = ", ".join("?" for _ in sheet.columns)
        con.executemany(f"INSERT INTO _sheet_buffer VALUES ({placeholders})", sheet.rows)
    con.execute(f"COPY _sheet_buffer TO '{out}' (FORMAT PARQUET)")
    con.execute("DROP TABLE _sheet_buffer")
    return out
