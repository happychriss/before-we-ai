#!/usr/bin/env python3
"""Pre-publication scan — does anything real-world sit in a public corpus?

Both landscapes assert they are fully fictional. This checks that with a tool
instead of trusting the assertion, and it is a CI gate so it stays checked.

What it can find, and what it cannot:

  CAN   structurally verifiable identities — e-mail addresses, IBANs (mod-97),
        German USt-IdNr and Polish NIP numbers whose *check digit is valid*
        (a made-up number almost never passes its own checksum), external URLs,
        and names from a denylist of large real companies.
  CANNOT decide whether a person's name is invented. "Marta Lewandowska,
        Procurement" is either a persona or a real employee, and no regex knows
        which. That judgement stays with the owner; this tool narrows the pile.

Every finding must end in one of two places: fixed, or written into
`scripts/publication-scan-waivers.yaml` with a reason. Silence is not an option
— an unwaived finding fails the build.

Deliberately not read: `expected_verdicts.yaml` and anything under an
`answer-key/` or `ground-truth/` directory. Those hold held-out traps; a tool
that prints matches from them would leak the answers into build logs.

    python scripts/publication-scan.py [--verbose]
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
WAIVERS = Path(__file__).resolve().parent / "publication-scan-waivers.yaml"

# Every landscape's business data and its domain guide — a new landscape is
# scanned the day it is added, with no list to remember to extend.
#
# Not the generators and not the answer keys. Not because they are safe: a
# generator is the *source* of every name in the data, so anything real in one
# arrives in data/ and is caught there. It is because those files are the ones
# nobody may read, and a tool that opens them and prints matched fragments into
# a build log is a way for the key to escape.
SCAN_ROOTS = sorted(
    [d for landscape in (REPO / "corpora").glob("*/")
     for d in (landscape / "data", landscape / "guide") if d.exists()])

# Never opened: held-out answers.
EXCLUDED_NAMES = {"expected_verdicts.yaml"}
EXCLUDED_DIRS = {"answer-key", "ground-truth", "generator", "spec", "__pycache__"}

TEXT_SUFFIXES = {".csv", ".txt", ".md", ".yaml", ".yml", ".json", ".sql", ".tsv"}


@dataclass(frozen=True)
class Finding:
    kind: str
    value: str
    where: str

    def key(self) -> str:
        return f"{self.kind}:{self.value}"


# ---------------------------------------------------------------- checksums

def _iban_valid(iban: str) -> bool:
    s = iban.replace(" ", "").upper()
    if not 15 <= len(s) <= 34:
        return False
    rearranged = s[4:] + s[:4]
    digits = "".join(str(ord(c) - 55) if c.isalpha() else c for c in rearranged)
    if not digits.isdigit():
        return False
    return int(digits) % 97 == 1


def _ustid_de_valid(digits: str) -> bool:
    """German USt-IdNr check digit (the ISO 7064-style scheme the BZSt uses)."""
    if len(digits) != 9 or not digits.isdigit():
        return False
    p, s = 10, 0
    for c in digits[:8]:
        s = (int(c) + p) % 10 or 10
        p = (2 * s) % 11
    check = (11 - p) % 10
    return check == int(digits[8])


def _nip_pl_valid(digits: str) -> bool:
    """Polish NIP weighted checksum."""
    if len(digits) != 10 or not digits.isdigit():
        return False
    weights = (6, 5, 7, 2, 3, 4, 5, 6, 7)
    total = sum(w * int(d) for w, d in zip(weights, digits))
    return total % 11 == int(digits[9])


# ---------------------------------------------------------------- patterns

EMAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
IBAN = re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{11,30}\b")
VAT_DE = re.compile(r"\bDE\s?(\d{9})\b")
VAT_PL = re.compile(r"\bPL\s?(\d{10})\b")
URL = re.compile(r"https?://([A-Za-z0-9.-]+)")

# Reserved by RFC 2606 / RFC 6761 for documentation and testing. An address in
# one of these cannot belong to anybody, so it is clean by construction rather
# than clean by waiver — there is nothing for the owner to judge.
RESERVED_DOMAINS = (".example", ".invalid", ".test", ".localhost",
                    "example.com", "example.org", "example.net")

# Large real companies whose name in a "fully fictional" corpus means a real
# document got mixed in. Not exhaustive by design — a denylist finds the
# obvious case, it does not certify the rest.
REAL_COMPANIES = [
    "bosch", "siemens", "sap se", "volkswagen", "daimler", "mercedes-benz",
    "bmw ag", "allianz", "deutsche bank", "commerzbank", "thyssenkrupp",
    "basf", "bayer ag", "lufthansa", "maersk", "hapag-lloyd", "meyer werft",
    "thyssen", "salesforce", "oracle corporation", "microsoft corporation",
]


def scan_text(text: str, where: str) -> list[Finding]:
    out: list[Finding] = []
    for m in EMAIL.finditer(text):
        domain = m.group(0).rsplit("@", 1)[1].lower()
        if domain.endswith(RESERVED_DOMAINS):
            continue
        out.append(Finding("email", m.group(0), where))
    for m in IBAN.finditer(text):
        if _iban_valid(m.group(0)):
            out.append(Finding("iban", m.group(0), where))
    for m in VAT_DE.finditer(text):
        if _ustid_de_valid(m.group(1)):
            out.append(Finding("vat_de", "DE" + m.group(1), where))
    for m in VAT_PL.finditer(text):
        if _nip_pl_valid(m.group(1)):
            out.append(Finding("vat_pl", "PL" + m.group(1), where))
    for m in URL.finditer(text):
        out.append(Finding("url", m.group(1).lower(), where))
    low = text.lower()
    for name in REAL_COMPANIES:
        if name in low:
            out.append(Finding("real_company", name, where))
    return out


# ---------------------------------------------------------------- readers

def read_any(path: Path) -> str | None:
    suffix = path.suffix.lower()
    try:
        if suffix in TEXT_SUFFIXES:
            return path.read_text(encoding="utf-8", errors="replace")
        if suffix == ".pdf":
            import pymupdf
            with pymupdf.open(path) as doc:
                return "\n".join(page.get_text() for page in doc)
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            book = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet in book.worksheets:
                parts.append(sheet.title)
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" ".join(str(v) for v in row if v is not None))
            book.close()
            return "\n".join(parts)
        if suffix == ".duckdb":
            import duckdb
            con = duckdb.connect(str(path), read_only=True)
            parts = []
            tables = [r[0] for r in con.execute(
                "SELECT table_name FROM information_schema.tables").fetchall()]
            for table in tables:
                parts.append(table)
                cols = con.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = ? AND data_type IN ('VARCHAR','TEXT')",
                    [table]).fetchall()
                parts += [c[0] for c in cols]
                for (col,) in cols:
                    rows = con.execute(
                        f'SELECT DISTINCT "{col}" FROM "{table}" '
                        f'WHERE "{col}" IS NOT NULL LIMIT 5000').fetchall()
                    parts += [str(r[0]) for r in rows]
            con.close()
            return "\n".join(parts)
    except Exception as exc:  # a reader that dies must not pass silently
        return f"<<UNREADABLE: {exc}>>"
    return None


def load_waivers() -> dict[str, str]:
    if not WAIVERS.exists():
        return {}
    import yaml
    loaded = yaml.safe_load(WAIVERS.read_text(encoding="utf-8")) or {}
    return {str(k): str(v) for k, v in (loaded.get("waived") or {}).items()}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verbose", action="store_true",
                        help="list every file scanned")
    args = parser.parse_args()

    waivers = load_waivers()
    findings: list[Finding] = []
    scanned = skipped = 0

    for root in SCAN_ROOTS:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*")):
            if not path.is_file():
                continue
            if path.name in EXCLUDED_NAMES or EXCLUDED_DIRS & set(path.parts):
                skipped += 1
                continue
            text = read_any(path)
            if text is None:
                if args.verbose:
                    print(f"  (no reader) {path.relative_to(REPO)}")
                continue
            rel = str(path.relative_to(REPO))
            if text.startswith("<<UNREADABLE:"):
                # A file nobody could open is a file nobody checked. That is a
                # finding, not a shrug.
                findings.append(Finding("unreadable", path.name, rel))
                continue
            scanned += 1
            if args.verbose:
                print(f"  scanned     {rel}")
            findings += scan_text(text, rel)

    unique: dict[str, Finding] = {}
    places: dict[str, set[str]] = {}
    for f in findings:
        unique.setdefault(f.key(), f)
        places.setdefault(f.key(), set()).add(f.where)

    unwaived = {k: v for k, v in unique.items() if k not in waivers}

    print(f"\npublication scan: {scanned} files read, {skipped} held-out files "
          f"skipped, {len(unique)} distinct findings "
          f"({len(waivers)} waived, {len(unwaived)} not)")

    if waivers and args.verbose:
        print("\nwaived:")
        for key, reason in sorted(waivers.items()):
            print(f"  {key}\n      {reason}")

    if not unwaived:
        print("clean.\n")
        return 0

    print("\nUNWAIVED — fix the corpus or add a reason to "
          f"{WAIVERS.relative_to(REPO)}:\n")
    for key in sorted(unwaived):
        where = ", ".join(sorted(places[key])[:4])
        print(f"  {key}\n      in: {where}")
    print()
    return 1


if __name__ == "__main__":
    sys.exit(main())
