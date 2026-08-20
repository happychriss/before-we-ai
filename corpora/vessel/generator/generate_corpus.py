#!/usr/bin/env python3
"""Generate a deliberately imperfect vessel-company test corpus.

The corpus is independent test data.  It models a fictional group and plants
cross-document contradictions, missing links, duplicates, and policy ambiguity.
Nothing in this generator imports or inspects the before-we-ai product.
"""

from __future__ import annotations

import hashlib
import json
import math
import random
import shutil
import textwrap
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
SOURCE_DIR = ROOT / "source-documents"
TRUTH_DIR = ROOT / "ground-truth"
SEED = 20250802
RNG = random.Random(SEED)
FIXED_TIME = datetime(2026, 1, 15, 9, 30, tzinfo=timezone.utc)

COMPANY = "Hanseatic Vessel Works Group GmbH"
SHORT = "HVW Group"

VESSELS = [
    {
        "project_id": "VSL-2407",
        "yard_alias": "HH-407",
        "contract_alias": "NB-407",
        "name": "M/V Nordlicht",
        "type": "coastal research vessel",
        "customer_id": "CUS-NO-014",
        "customer": "Nordsee Forschung AS",
        "yard": "Hamburg",
        "contract_value": 3_400_000,
        "status": "delivered",
        "acceptance": date(2025, 3, 28),
        "true_cost": 2_718_400,
        "forecast_cost": 2_718_400,
    },
    {
        "project_id": "VSL-2411",
        "yard_alias": "GD-11/B",
        "contract_alias": "BALTIC-11",
        "name": "Baltic Surveyor",
        "type": "offshore survey vessel",
        "customer_id": "CUS-DK-022",
        "customer": "Skagerrak Marine ApS",
        "yard": "Gdansk",
        "contract_value": 4_000_000,
        "status": "delivered",
        "acceptance": date(2025, 9, 19),
        "true_cost": 3_286_900,
        "forecast_cost": 3_286_900,
    },
    {
        "project_id": "VSL-2502",
        "yard_alias": "ER-02",
        "contract_alias": "ELBE-RUNNER",
        "name": "Elbe Runner",
        "type": "hybrid crew-transfer vessel",
        "customer_id": "CUS-DE-031",
        "customer": "Elbe Wind Service GmbH",
        "yard": "Gdansk",
        "contract_value": 2_350_000,
        "status": "work in progress",
        "acceptance": None,
        "true_cost": 1_563_200,
        "forecast_cost": 2_204_000,
    },
    {
        "project_id": "VSL-2504",
        "yard_alias": "AMBER-77",
        "contract_alias": "AP-77",
        "name": "Amber Pilot",
        "type": "harbour pilot boat",
        "customer_id": "CUS-PL-008",
        "customer": "Port Pilot Gdansk S.A.",
        "yard": "Gdansk",
        "contract_value": 1_480_000,
        "status": "delivered",
        "acceptance": date(2025, 12, 12),
        "true_cost": 1_214_600,
        "forecast_cost": 1_214_600,
    },
    {
        "project_id": "VSL-2506",
        "yard_alias": "TUG-X6",
        "contract_alias": "LOI-6",
        "name": "Hansa Tug",
        "type": "harbour tug concept",
        "customer_id": "CUS-SE-041",
        "customer": "Oresund Towage AB",
        "yard": "Hamburg",
        "contract_value": 3_100_000,
        "status": "cancelled letter of intent",
        "acceptance": None,
        "true_cost": 84_500,
        "forecast_cost": 0,
    },
]

TRUE_COSTS = {
    "VSL-2407": {
        "steel and hull": 650_000,
        "engine package": 520_000,
        "propulsion": 180_000,
        "navigation": 165_000,
        "electrical": 205_000,
        "outfitting": 280_000,
        "direct labour": 360_000,
        "yard overhead": 210_000,
        "design": 88_000,
        "classification": 42_000,
        "logistics": 18_400,
    },
    "VSL-2411": {
        "steel and hull": 700_000,
        "engine package": 650_000,
        "propulsion": 210_000,
        "navigation": 190_000,
        "electrical": 245_000,
        "outfitting": 510_000,
        "direct labour": 380_000,
        "yard overhead": 220_000,
        "design": 105_000,
        "classification": 48_000,
        "logistics": 28_900,
    },
    "VSL-2502": {
        "steel and hull": 480_000,
        "engine package": 400_000,
        "propulsion": 70_000,
        "navigation": 35_000,
        "electrical": 90_000,
        "outfitting": 130_000,
        "direct labour": 190_000,
        "yard overhead": 95_000,
        "design": 55_000,
        "classification": 18_000,
        "logistics": 200,
    },
    "VSL-2504": {
        "steel and hull": 240_000,
        "engine package": 270_000,
        "propulsion": 85_000,
        "navigation": 72_000,
        "electrical": 95_000,
        "outfitting": 140_000,
        "direct labour": 145_000,
        "yard overhead": 90_000,
        "design": 40_000,
        "classification": 25_000,
        "logistics": 12_600,
    },
    "VSL-2506": {
        "design": 62_000,
        "sales engineering": 22_500,
    },
}

TRUE_EXTERNAL_REVENUE_2025 = 11_460_000
INTERNAL_POC_REVENUE = 1_598_000
MANAGEMENT_REVENUE_2025 = TRUE_EXTERNAL_REVENUE_2025 + INTERNAL_POC_REVENUE

NAVY = "19324a"
BLUE = "2f75b5"
LIGHT_BLUE = "d9eaf7"
GOLD = "d6a847"
LIGHT_GOLD = "f6ecd3"
RED = "c74242"
LIGHT_RED = "f9dddd"
GREEN = "3f7d59"
GREY = "6b7280"
LIGHT_GREY = "eef1f4"
WHITE = "ffffff"

THIN = Side(style="thin", color="c7cdd4")


def clean_output() -> None:
    for directory in (SOURCE_DIR, TRUTH_DIR):
        if directory.exists():
            shutil.rmtree(directory)
        directory.mkdir(parents=True)


def new_workbook(title: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "Hanseatic Vessel Works Group GmbH"
    wb.properties.title = title
    wb.properties.subject = "Fictional vessel-company test corpus"
    wb.properties.description = (
        "Synthetic business document. Contains deliberate inconsistencies."
    )
    wb.properties.created = FIXED_TIME.replace(tzinfo=None)
    wb.properties.modified = FIXED_TIME.replace(tzinfo=None)
    return wb


def style_sheet(ws, freeze: str = "A2", auto_filter: bool = True) -> None:
    ws.freeze_panes = freeze
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN)
        ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="dfe3e8"))
    for col_idx, column in enumerate(ws.iter_cols(), start=1):
        values = [str(c.value) if c.value is not None else "" for c in column[:80]]
        width = min(42, max(10, max((len(v) for v in values), default=8) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if auto_filter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_rows(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def money_format(ws, columns: list[int]) -> None:
    for col in columns:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00;[Red]-#,##0.00;"-"'


def create_master_workbook() -> None:
    wb = new_workbook("Company, customer and vessel master")

    ws = wb.create_sheet("Group_Structure")
    rows = [
        ["HVW-DE", COMPANY, "GmbH", "Hamburg", "Germany", "EUR", "DE291845771", True, "parent", "100% of BHW-PL"],
        ["BHW-PL", "Baltic Hull Works sp. z o.o.", "sp. z o.o.", "Gdansk", "Poland", "PLN", "PL5842991037", False, "subsidiary", "Wholly owned by HVW-DE"],
        ["GD-YARD", "Baltic Hull Works - Danzig branch", "branch", "Danzig", "Poland", "PLN", "PL5842991037", False, "operating location", "Old label; not a separate legal entity"],
    ]
    write_rows(ws, ["entity_code", "legal_name", "legal_form", "city", "country", "functional_currency", "vat_id_fictional", "headquarter", "relationship", "comment"], rows)
    style_sheet(ws)
    add_table(ws, "GroupStructure")

    ws = wb.create_sheet("Vessel_Master")
    rows = []
    for vessel in VESSELS:
        rows.append([
            vessel["project_id"], vessel["yard_alias"], vessel["contract_alias"],
            vessel["name"], vessel["type"], vessel["customer_id"], vessel["customer"],
            vessel["yard"], vessel["status"], vessel["acceptance"], vessel["contract_value"],
            "EUR", vessel["forecast_cost"],
            "customer acceptance" if vessel["project_id"] != "VSL-2502" else "management says cost-to-cost; contract says acceptance",
        ])
    write_rows(ws, ["project_id", "yard_job_no", "contract_reference", "vessel_name", "vessel_type", "customer_id", "customer_name", "build_yard", "status", "acceptance_date", "contract_value", "contract_currency", "latest_forecast_cost", "revenue_recognition_note"], rows)
    style_sheet(ws)
    money_format(ws, [11, 13])
    add_table(ws, "VesselMaster")

    ws = wb.create_sheet("Customers")
    customer_rows = [
        ["CUS-NO-014", "Nordsee Forschung AS", "NO", "external", "EUR", "Nordsee Research"],
        ["CUS-DK-022", "Skagerrak Marine ApS", "DK", "external", "EUR", "Skagerrak Marine A/S"],
        ["CUS-DE-031", "Elbe Wind Service GmbH", "DE", "external", "EUR", "ElbeWind"],
        ["CUS-PL-008", "Port Pilot Gdansk S.A.", "PL", "external", "EUR", "Port Pilot Danzig"],
        ["CUS-SE-041", "Oresund Towage AB", "SE", "external", "EUR", "Oresund Towing"],
        ["IC-BHW-PL", "Baltic Hull Works sp. z o.o.", "PL", "intercompany", "EUR", "Danzig Yard"],
        ["IC-HVW-DE", COMPANY, "DE", "intercompany", "EUR", "Hamburg HQ"],
        ["CUS-UK-017", "North Channel Ferries Ltd", "GB", "external", "EUR", "NCF"],
        ["CUS-DE-044", "Helgoland Service KG", "DE", "external", "EUR", "Helgoland"],
    ]
    write_rows(ws, ["customer_id", "legal_name", "country", "relationship", "billing_currency", "common_alias"], customer_rows)
    style_sheet(ws)
    add_table(ws, "CustomerMaster")

    ws = wb.create_sheet("Project_Aliases")
    alias_rows = [
        ["VSL-2407", "HH-407", "NB-407", "Nordlicht", "NORD-407", "approved"],
        ["VSL-2411", "GD-11/B", "BALTIC-11", "Baltic Surveyor", "BS-11", "approved"],
        ["VSL-2502", "ER-02", "ELBE-RUNNER", "Elbe Runner", "EWR-02", "approved"],
        ["VSL-2504", "AMBER-77", "AP-77", "Amber Pilot", "AP77", "approved"],
        ["VSL-2506", "TUG-X6", "LOI-6", "Hansa Tug", "", "LOI only"],
        ["", "Danzig-11", "", "Baltic Surveyor", "", "unconfirmed old yard code"],
    ]
    write_rows(ws, ["project_id", "yard_job_no", "contract_reference", "vessel_name", "finance_alias", "mapping_status"], alias_rows)
    style_sheet(ws)
    add_table(ws, "ProjectAliases")
    ws.sheet_state = "hidden"

    ws = wb.create_sheet("Read_Me_First", 0)
    ws["A1"] = "MASTER DATA EXPORT"
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")
    notes = [
        "Export owner: Group Controlling, Hamburg",
        "Reporting date: 31 December 2025",
        "Warning: Danzig and Gdansk are both used in legacy systems.",
        "Warning: yard job numbers and contract references are not project IDs.",
        "The Project_Aliases sheet is hidden because it was marked 'work in progress'.",
        "VAT IDs and every company/customer in this corpus are fictional.",
    ]
    for idx, note in enumerate(notes, start=3):
        ws.cell(idx, 1, note)
    ws.column_dimensions["A"].width = 95
    wb.save(SOURCE_DIR / "01_company_and_vessel_master.xlsx")


def split_amount(total: int, count: int) -> list[int]:
    if count == 1:
        return [total]
    weights = [RNG.randint(10, 100) for _ in range(count)]
    raw = [round(total * w / sum(weights)) for w in weights[:-1]]
    raw.append(total - sum(raw))
    return raw


def create_cost_ledger() -> dict[str, float]:
    wb = new_workbook("Project cost ledger 2025")
    ws = wb.create_sheet("Cost_Ledger")
    headers = [
        "transaction_id", "posting_date", "entity", "location", "project_id",
        "yard_job_ref", "category", "vendor", "invoice_number", "po_number",
        "description", "currency", "amount_local", "pln_per_eur", "reported_eur",
        "vat_code", "approved", "source_system",
    ]
    rows: list[list] = []
    vendors = {
        "steel and hull": ["Baltic Steel S.A.", "Nordblech GmbH", "Stal-Marine sp. z o.o."],
        "engine package": ["HanseDiesel GmbH", "Baltic Engines Sp. z o.o."],
        "propulsion": ["PropTech Marine BV", "Baltic Propulsion S.A."],
        "navigation": ["North Sea Electronics GmbH", "Navico Works Sp. z o.o."],
        "electrical": ["Elbe Cable Systems GmbH", "Gdansk Electrics S.A."],
        "outfitting": ["Marine Interior GmbH", "Danzig Outfitters Sp. z o.o."],
        "direct labour": ["Internal payroll allocation"],
        "yard overhead": ["Internal yard allocation"],
        "design": ["HVW Design Office", "Baltic Naval Architects"],
        "classification": ["North Atlantic Register"],
        "logistics": ["Harbour Logistics KG"],
        "sales engineering": ["HVW Bid Office"],
    }
    vessel_map = {v["project_id"]: v for v in VESSELS}
    txn = 1
    date_base = date(2024, 2, 1)
    for project_id, categories in TRUE_COSTS.items():
        vessel = vessel_map[project_id]
        for category, total in categories.items():
            count = 1 if total < 25_000 else RNG.randint(2, 5)
            for part in split_amount(total, count):
                posting = date_base + timedelta(days=RNG.randint(0, 680))
                location = vessel["yard"]
                entity = "BHW-PL" if location == "Gdansk" else "HVW-DE"
                use_pln = location == "Gdansk" and category not in {"engine package", "propulsion", "classification"} and RNG.random() < 0.72
                fx = round(RNG.uniform(4.21, 4.42), 4)
                currency = "PLN" if use_pln else "EUR"
                local = round(part * fx, 2) if use_pln else float(part)
                reported_eur = float(part)
                vendor = RNG.choice(vendors[category])
                invoice_no = f"{vendor[:3].upper().replace(' ', '')}-{posting.year}-{RNG.randint(1000,9999)}"
                row = [
                    f"CST-{txn:05d}", posting, entity, location, project_id,
                    vessel["yard_alias"], category, vendor, invoice_no,
                    f"PO-{posting.year}-{RNG.randint(10000,99999)}",
                    f"{category.title()} for {vessel['name']}", currency, local,
                    fx if use_pln else 1.0, reported_eur,
                    "PL23" if entity == "BHW-PL" else "DE19", True,
                    "YARD-PL" if entity == "BHW-PL" else "ERP-DE",
                ]
                rows.append(row)
                txn += 1

    traps: dict[str, float] = {}

    # Blank project: a completed-vessel safety/classification cost can only be
    # recovered from the PO reference and the operations notes PDF.
    blank_idx = next(i for i, r in enumerate(rows) if r[4] == "VSL-2407" and r[6] == "classification")
    traps["blank_project_cost"] = float(rows[blank_idx][14])
    rows[blank_idx][4] = ""
    rows[blank_idx][10] = "Safety and class attendance for NB-407; project field rejected by ERP"

    # Misallocation: an Amber Pilot engine instalment is booked to Elbe Runner.
    wrong_idx = next(i for i, r in enumerate(rows) if r[4] == "VSL-2504" and r[6] == "engine package")
    traps["misallocated_engine"] = float(rows[wrong_idx][14])
    rows[wrong_idx][4] = "VSL-2502"
    rows[wrong_idx][10] = "Main engine instalment for AMBER-77 (AP-77 acceptance package)"

    # Duplicate invoice: same invoice, amount, PO and description, different transaction id.
    dup_idx = next(i for i, r in enumerate(rows) if r[4] == "VSL-2411" and r[6] == "steel and hull")
    duplicate = list(rows[dup_idx])
    duplicate[0] = f"CST-{txn:05d}"
    duplicate[1] = duplicate[1] + timedelta(days=2)
    duplicate[17] = "MANUAL-JOURNAL"
    traps["duplicate_supplier_invoice"] = float(duplicate[14])
    rows.append(duplicate)
    txn += 1

    # FX inversion/multiplication: local amount and rate are usable, reported EUR is absurd.
    fx_idx = next(i for i, r in enumerate(rows) if r[4] == "VSL-2502" and r[11] == "PLN" and r[14] > 20_000)
    correct_eur = float(rows[fx_idx][14])
    rows[fx_idx][14] = round(float(rows[fx_idx][12]) * float(rows[fx_idx][13]), 2)
    rows[fx_idx][10] += " / ERP conversion batch FX-UPLOAD-08"
    traps["fx_error_correct_eur"] = correct_eur
    traps["fx_error_reported_eur"] = float(rows[fx_idx][14])

    # Missing December labour for Elbe Runner: removed from the ledger, present in time records.
    labour_idxs = [i for i, r in enumerate(rows) if r[4] == "VSL-2502" and r[6] == "direct labour"]
    missing_idx = labour_idxs[-1]
    traps["missing_december_labour"] = float(rows[missing_idx][14])
    rows.pop(missing_idx)

    # Unapproved change order costs are real costs but not contract revenue.
    rows.append([
        f"CST-{txn:05d}", date(2025, 8, 31), "BHW-PL", "Gdansk", "VSL-2411",
        "GD-11/B", "outfitting", "Danzig Outfitters Sp. z o.o.", "DO-CO17-884",
        "PO-2025-88417", "Luxury cabin upgrade requested verbally; CO-17 not signed",
        "EUR", 90_000.0, 1.0, 90_000.0, "PL23", False, "YARD-PL",
    ])
    traps["unapproved_change_cost"] = 90_000.0

    write_rows(ws, headers, rows)
    style_sheet(ws)
    money_format(ws, [13, 15])
    add_table(ws, "CostLedger")
    for row in ws.iter_rows(min_row=2):
        if not row[4].value or row[16].value is False:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GOLD)

    ws_sum = wb.create_sheet("Pivot_Export")
    ws_sum.append(["project_id", "reported_eur_total", "note"])
    projects = [v["project_id"] for v in VESSELS]
    for idx, project in enumerate(projects, start=2):
        ws_sum.cell(idx, 1, project)
        ws_sum.cell(idx, 2, f'=SUMIF(Cost_Ledger!E:E,A{idx},Cost_Ledger!O:O)')
        ws_sum.cell(idx, 3, "Formula trusts project_id and reported_eur without quality checks")
    ws_sum.cell(len(projects) + 3, 1, "UNASSIGNED")
    ws_sum.cell(len(projects) + 3, 2, '=SUMIF(Cost_Ledger!E:E,"",Cost_Ledger!O:O)')
    style_sheet(ws_sum)
    money_format(ws_sum, [2])

    ws_codes = wb.create_sheet("Cost_Centre_Map")
    mappings = [["100-HH", "Hamburg production", "HVW-DE"], ["200-GD", "Gdansk production", "BHW-PL"], ["210-DZ", "Danzig old yard code", "BHW-PL"], ["900-HQ", "Hamburg headquarters", "HVW-DE"]]
    write_rows(ws_codes, ["cost_centre", "description", "entity"], mappings)
    style_sheet(ws_codes)

    wb.save(SOURCE_DIR / "02_project_cost_ledger_2025.xlsx")
    return traps


def create_time_and_overhead(traps: dict[str, float]) -> None:
    wb = new_workbook("Time and overhead allocations")
    ws = wb.create_sheet("Time_Allocations")
    headers = ["entry_id", "month", "employee_id", "home_location", "project_reference", "work_package", "hours", "hourly_rate_eur", "extended_cost_eur", "approval_status", "comment"]
    rows = []
    entry = 1
    targets = {
        "VSL-2407": 360_000,
        "VSL-2411": 380_000,
        "VSL-2502": 190_000,
        "VSL-2504": 145_000,
        "VSL-2506": 62_000,
    }
    aliases = {v["project_id"]: v["yard_alias"] for v in VESSELS}
    for project, target in targets.items():
        months = 8 if project in {"VSL-2407", "VSL-2411"} else 10
        amounts = split_amount(target, months)
        for month_idx, amount in enumerate(amounts, start=1):
            rate = RNG.choice([42.5, 48.0, 55.0, 62.0, 78.0])
            hours = round(amount / rate, 2)
            home = "Gdansk" if project in {"VSL-2411", "VSL-2502", "VSL-2504"} else "Hamburg"
            ref = aliases[project] if month_idx % 3 == 0 else project
            rows.append([
                f"TS-{entry:05d}", date(2025, min(month_idx, 12), 1),
                f"EMP-{RNG.randint(1,68):03d}", home, ref,
                RNG.choice(["hull", "electrical", "outfitting", "design", "commissioning"]),
                hours, rate, float(amount), "approved", "",
            ])
            entry += 1

    # The ledger omitted this approved December batch. It is included in the
    # 190k true cost by replacing one generic Elbe allocation with a specific line.
    elbe_rows = [i for i, r in enumerate(rows) if r[4] in {"VSL-2502", "ER-02"}]
    last = elbe_rows[-1]
    original = float(rows[last][8])
    rows[last][1] = date(2025, 12, 1)
    rows[last][2] = "BATCH-GD-DEC"
    rows[last][5] = "December yard close"
    rows[last][10] = "Approved 8 Jan 2026; interface to project ledger failed"
    traps["time_sheet_december_batch"] = original

    # One Hamburg allocation is duplicated in the export.
    dup_idx = next(i for i, r in enumerate(rows) if r[4] in {"VSL-2407", "HH-407"} and r[8] > 30_000)
    dup = list(rows[dup_idx])
    dup[0] = f"TS-{entry:05d}"
    dup[9] = "approved-copy"
    dup[10] = "Copied from offline supervisor workbook"
    traps["duplicate_timesheet"] = float(dup[8])
    rows.append(dup)

    write_rows(ws, headers, rows)
    style_sheet(ws)
    money_format(ws, [8, 9])
    add_table(ws, "TimeAllocations")

    ws = wb.create_sheet("Overhead_Pools")
    pools = [
        ["2025", "Hamburg", "production overhead", 780_000, "direct labour hours", 19_820, 39.35, "Includes HQ occupancy? controller says no"],
        ["2025", "Gdansk", "production overhead", 1_150_000, "machine hours", 25_200, 45.63, "Board report uses labour hours instead"],
        ["2025", "Hamburg HQ", "administration", 920_000, "revenue", 13_100_000, 0.0702, "Not normally inventoried; one project sheet includes it"],
    ]
    write_rows(ws, ["year", "location", "pool", "pool_eur", "allocation_base", "base_quantity", "rate", "comment"], pools)
    style_sheet(ws)
    money_format(ws, [4, 6, 7])
    add_table(ws, "OverheadPools")

    ws = wb.create_sheet("Employee_Rates")
    rates = []
    for emp in range(1, 69):
        rates.append([f"EMP-{emp:03d}", "Hamburg" if emp <= 23 else "Gdansk", RNG.choice([42.5, 48.0, 55.0, 62.0, 78.0]), "EUR", "2025-01-01", "Some Gdansk rates translated from PLN at budget rate"])
    write_rows(ws, ["employee_id", "location", "burdened_rate", "currency", "valid_from", "note"], rates)
    style_sheet(ws)
    add_table(ws, "EmployeeRates")

    wb.save(SOURCE_DIR / "03_time_and_overhead_allocations.xlsx")


def invoice_row(doc_id, doc_type, inv_date, entity, customer_id, project_ref, description, currency, net, vat_rate, status, accounting_treatment, due_date=None):
    sign = -1 if doc_type == "credit_note" else 1
    vat = round(net * vat_rate, 2)
    gross = round(net + vat, 2)
    return [doc_id, doc_type, inv_date, due_date or inv_date + timedelta(days=30), entity, customer_id, project_ref, description, currency, net, vat_rate, vat, gross, status, accounting_treatment, sign]


def create_sales_workbook() -> None:
    wb = new_workbook("Sales invoices and credit notes")
    ws = wb.create_sheet("Documents")
    rows = []
    # Vessel milestones. Revenue belongs to acceptance year, while invoicing spans years.
    rows += [
        invoice_row("DE-2024-1187", "invoice", date(2024, 5, 15), "HVW-DE", "CUS-NO-014", "NB-407", "Nordlicht contract deposit", "EUR", 1_020_000, 0.00, "posted", "contract liability until acceptance"),
        invoice_row("DE-2025-0211", "invoice", date(2025, 1, 31), "HVW-DE", "CUS-NO-014", "HH-407", "Nordlicht launch milestone", "EUR", 1_020_000, 0.00, "posted", "contract liability until acceptance"),
        invoice_row("DE-2025-0478", "invoice", date(2025, 3, 28), "HVW-DE", "CUS-NO-014", "VSL-2407", "Nordlicht acceptance and retention", "EUR", 1_360_000, 0.00, "posted", "revenue on customer acceptance"),
        invoice_row("PL/2024/00881", "invoice", date(2024, 11, 20), "BHW-PL", "CUS-DK-022", "BALTIC-11", "Baltic Surveyor keel milestone", "EUR", 800_000, 0.00, "posted", "contract liability until acceptance"),
        invoice_row("PL/2025/00317", "invoice", date(2025, 4, 30), "BHW-PL", "CUS-DK-022", "GD-11/B", "Baltic Surveyor machinery milestone", "EUR", 1_200_000, 0.00, "posted", "contract liability until acceptance"),
        invoice_row("PL/2025/00642", "invoice", date(2025, 7, 31), "BHW-PL", "CUS-DK-022", "Danzig-11", "Baltic Surveyor sea-trial milestone", "EUR", 1_200_000, 0.00, "posted", "contract liability until acceptance"),
        invoice_row("PL/2025/00813", "invoice", date(2025, 9, 19), "BHW-PL", "CUS-DK-022", "VSL-2411", "Baltic Surveyor acceptance", "EUR", 800_000, 0.00, "posted", "revenue on customer acceptance"),
        invoice_row("PL/2025/00599", "invoice", date(2025, 6, 30), "BHW-PL", "CUS-PL-008", "AP-77", "Amber Pilot steel and engine", "EUR", 592_000, 0.23, "posted", "contract liability until acceptance"),
        invoice_row("PL/2025/00901", "invoice", date(2025, 10, 31), "BHW-PL", "CUS-PL-008", "AMBER-77", "Amber Pilot launch", "EUR", 592_000, 0.22, "posted", "contract liability; VAT rate copied from old template"),
        invoice_row("PL/2025/01088", "invoice", date(2025, 12, 12), "BHW-PL", "CUS-PL-008", "VSL-2504", "Amber Pilot acceptance", "EUR", 296_000, 0.23, "posted", "revenue on customer acceptance"),
        invoice_row("DE-2025-1032", "invoice", date(2025, 5, 31), "HVW-DE", "CUS-DE-031", "ELBE-RUNNER", "Elbe Runner contract deposit", "EUR", 705_000, 0.19, "posted", "customer advance; no revenue before acceptance"),
        invoice_row("PL/2025/00977", "invoice", date(2025, 11, 30), "BHW-PL", "CUS-DE-031", "ER-02", "Elbe Runner hull milestone", "EUR", 470_000, 0.00, "posted", "customer advance; intra-EU treatment pending evidence"),
        invoice_row("DE-2025-0777", "invoice", date(2025, 3, 10), "HVW-DE", "CUS-SE-041", "TUG-X6", "Refundable reservation deposit for Hansa Tug", "EUR", 310_000, 0.00, "posted", "refundable liability, not revenue"),
        # Positive credit-note amount: the document type supplies the sign.
        invoice_row("DE-CN-2025-0777", "credit_note", date(2025, 6, 2), "HVW-DE", "CUS-SE-041", "LOI-6", "Cancellation credit for Hansa Tug reservation", "EUR", 310_000, 0.00, "posted", "reduces refundable liability"),
    ]
    # External service revenue: 2.58m, of which only 200k is linked to a built vessel.
    rows += [
        invoice_row("DE-2025-1310", "invoice", date(2025, 7, 15), "HVW-DE", "CUS-UK-017", "REFIT-NCF-1", "Ferry refit package", "EUR", 920_000, 0.00, "posted", "service revenue over completed work packages"),
        invoice_row("PL/2025/00771", "invoice", date(2025, 8, 20), "BHW-PL", "CUS-DE-044", "REPAIR-HLG", "Hull repair and dry-dock", "EUR", 620_000, 0.00, "posted", "service revenue on completion"),
        invoice_row("DE-2025-1442", "invoice", date(2025, 10, 9), "HVW-DE", "CUS-NO-014", "VSL-2407", "Post-delivery science-module upgrade", "EUR", 200_000, 0.00, "posted", "service revenue"),
        invoice_row("DE-2025-1511", "invoice", date(2025, 11, 3), "HVW-DE", "CUS-DE-044", "", "Spare parts and mobile service teams", "EUR", 340_000, 0.19, "posted", "parts and service revenue"),
        invoice_row("DE-2025-1588", "invoice", date(2025, 12, 18), "HVW-DE", "CUS-UK-017", "DESIGN-42", "Concept design licence", "EUR", 500_000, 0.00, "posted", "licence revenue; not a vessel build"),
    ]
    # Intercompany invoices should disappear on consolidation.
    rows += [
        invoice_row("IC-DE-25031", "invoice", date(2025, 3, 31), "HVW-DE", "IC-BHW-PL", "VSL-2411", "Engineering and procurement recharge Q1", "EUR", 315_000, 0.00, "posted", "intercompany revenue"),
        invoice_row("IC-DE-25062", "invoice", date(2025, 6, 30), "HVW-DE", "IC-BHW-PL", "VSL-2502", "Engineering and procurement recharge Q2", "EUR", 280_000, 0.00, "posted", "intercompany revenue"),
        invoice_row("IC-PL-25091", "invoice", date(2025, 9, 30), "BHW-PL", "IC-HVW-DE", "VSL-2407", "Yard labour recharge", "EUR", 365_000, 0.00, "posted", "intercompany revenue"),
        invoice_row("IC-PL-25121", "invoice", date(2025, 12, 31), "BHW-PL", "IC-HVW-DE", "VSL-2504", "Yard capacity and launch support", "EUR", 300_000, 0.00, "posted", "intercompany revenue"),
    ]
    # A draft duplicate shares a number with a posted Polish invoice.
    rows.append(invoice_row("PL/2025/00813", "invoice", date(2025, 9, 18), "BHW-PL", "CUS-DK-022", "BS-11", "Acceptance invoice preview - DO NOT POST", "EUR", 800_000, 0.00, "draft", "not posted"))

    headers = ["document_id", "document_type", "document_date", "due_date", "issuing_entity", "customer_id", "project_reference", "description", "currency", "net_amount", "vat_rate", "vat_amount", "gross_amount", "status", "accounting_treatment", "sign_hint"]
    write_rows(ws, headers, rows)
    style_sheet(ws)
    money_format(ws, [10, 12, 13])
    for cell in ws["K"][1:]:
        cell.number_format = "0%"
    add_table(ws, "SalesDocuments")
    for row in ws.iter_rows(min_row=2):
        if row[13].value == "draft" or row[10].value == 0.22:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GOLD)

    ws = wb.create_sheet("Management_Recognition")
    recognition_rows = [
        ["VSL-2407", "M/V Nordlicht", 3_400_000, "accepted 2025-03-28", "external vessel revenue"],
        ["VSL-2411", "Baltic Surveyor", 4_000_000, "accepted 2025-09-19", "external vessel revenue"],
        ["VSL-2504", "Amber Pilot", 1_480_000, "accepted 2025-12-12", "external vessel revenue"],
        ["VSL-2502", "Elbe Runner", INTERNAL_POC_REVENUE, "68% internal cost-to-cost estimate", "management-only POC; disputed by contract"],
        ["SERVICES", "Refit, repair, upgrade, spares, design", 2_580_000, "completed or delivered in 2025", "external non-vessel revenue"],
        ["VSL-2506", "Hansa Tug", 0, "LOI cancelled", "deposit is refundable"],
    ]
    write_rows(ws, ["project_id", "description", "revenue_2025_eur", "recognition_basis", "classification"], recognition_rows)
    style_sheet(ws)
    money_format(ws, [3])
    ws["A9"] = "Management total"
    ws["C9"] = "=SUM(C2:C7)"
    ws["A10"] = "External revenue excluding disputed POC"
    ws["C10"] = f"={TRUE_EXTERNAL_REVENUE_2025}"
    ws["A11"] = "Warning"
    ws["C11"] = "This sheet is a management schedule, not the audited revenue ledger."

    ws = wb.create_sheet("Revenue_By_Ship")
    ship_rows = [
        ["VSL-2407", "M/V Nordlicht", 3_400_000, 200_000, 3_600_000, "Includes post-delivery upgrade"],
        ["VSL-2411", "Baltic Surveyor", 4_000_000, 0, 4_000_000, "Contract sale only"],
        ["VSL-2504", "Amber Pilot", 1_480_000, 0, 1_480_000, "Contract sale only"],
        ["VSL-2502", "Elbe Runner", 0, 0, 0, "Not accepted; management separately books POC"],
        ["VSL-2506", "Hansa Tug", 0, 0, 0, "LOI cancelled"],
        ["OTHER", "Customer vessels not built by HVW", 0, 1_540_000, 1_540_000, "Refit and repair"],
        ["UNALLOCATED", "Spares and design licence", 0, 840_000, 840_000, "No reliable vessel identifier"],
    ]
    write_rows(ws, ["project_id", "vessel", "vessel_contract_revenue", "service_revenue", "total_linked_revenue", "note"], ship_rows)
    style_sheet(ws)
    money_format(ws, [3, 4, 5])

    chart = BarChart()
    chart.title = "Revenue linked to vessel/project"
    data = Reference(ws, min_col=5, min_row=1, max_row=6)
    cats = Reference(ws, min_col=2, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "H2")

    wb.save(SOURCE_DIR / "04_sales_invoices_and_credit_notes_2025.xlsx")


def create_intercompany_workbook() -> None:
    wb = new_workbook("Intercompany recharges and elimination")
    ws = wb.create_sheet("IC_Transactions")
    rows = [
        ["IC-DE-25031", date(2025, 3, 31), "HVW-DE", "BHW-PL", "VSL-2411", "engineering and procurement", 300_000, 0.05, 315_000, "EUR", "IC-PL-3103", 315_000, "matched"],
        ["IC-DE-25062", date(2025, 6, 30), "HVW-DE", "BHW-PL", "VSL-2502", "engineering and procurement", 260_465, 0.075, 280_000, "EUR", "IC-PL-6205", 278_600, "rate/rounding difference"],
        ["IC-PL-25091", date(2025, 9, 30), "BHW-PL", "HVW-DE", "VSL-2407", "yard labour", 347_619, 0.05, 365_000, "EUR", "IC-DE-1905", 365_000, "matched"],
        ["IC-PL-25121", date(2025, 12, 31), "BHW-PL", "HVW-DE", "VSL-2504", "launch support", 285_714, 0.05, 300_000, "EUR", "", 0, "missing receiving-side booking"],
        ["IC-DE-25122", date(2025, 12, 31), "HVW-DE", "BHW-PL", "GROUP", "HQ management fee", 120_000, 0.05, 126_000, "EUR", "IC-PL-2212", 126_000, "present only in IC workbook, absent sales export"],
    ]
    write_rows(ws, ["sender_document", "date", "seller_entity", "buyer_entity", "project_id", "service", "cost_base_eur", "markup_rate", "sender_amount_eur", "currency", "receiver_document", "receiver_amount_eur", "match_status"], rows)
    style_sheet(ws)
    money_format(ws, [7, 9, 12])
    for cell in ws["H"][1:]:
        cell.number_format = "0.0%"
    add_table(ws, "ICTransactions")

    ws = wb.create_sheet("Elimination_Journal")
    eliminations = [
        ["ELIM-001", "IC-DE-25031", -315_000, 315_000, "posted"],
        ["ELIM-002", "IC-DE-25062", -280_000, 278_600, "out of balance by 1,400"],
        ["ELIM-003", "IC-PL-25091", -365_000, 365_000, "posted"],
        ["ELIM-004", "IC-PL-25121", -300_000, 0, "receiver missing; not posted"],
        ["ELIM-005", "IC-DE-25122", 0, 0, "forgotten"],
    ]
    write_rows(ws, ["journal_id", "source_document", "eliminate_revenue_eur", "eliminate_cost_eur", "status"], eliminations)
    style_sheet(ws)
    money_format(ws, [3, 4])
    add_table(ws, "EliminationJournal")

    ws = wb.create_sheet("Policy")
    policy = [
        ["Engineering recharge", "cost plus 5%", "Group tax memo", "Controller workbook uses 7.5% for Q2 without approval"],
        ["Yard labour", "cost plus 5%", "Group tax memo", "Should eliminate on consolidation"],
        ["HQ management fee", "fixed quarterly allocation", "service agreement missing", "Tax reviewer requested evidence"],
    ]
    write_rows(ws, ["service", "policy", "authority", "exception"], policy)
    style_sheet(ws)

    wb.save(SOURCE_DIR / "05_intercompany_recharges_2025.xlsx")


def create_bank_and_fx_workbook() -> None:
    wb = new_workbook("Bank receipts, invoice applications and FX")
    ws = wb.create_sheet("Bank_Receipts")
    rows = [
        ["RCPT-250104", date(2025, 1, 4), "HVW-DE", "CUS-NO-014", "DE-2024-1187", 1_020_000, "EUR", 1.0, 1_020_000, "prior-year invoice"],
        ["RCPT-250228", date(2025, 2, 28), "HVW-DE", "CUS-NO-014", "DE-2025-0211", 1_020_000, "EUR", 1.0, 1_020_000, "progress billing"],
        ["RCPT-250415", date(2025, 4, 15), "HVW-DE", "CUS-NO-014", "DE-2025-0478", 1_240_000, "EUR", 1.0, 1_240_000, "120k retention unpaid at year end"],
        ["RCPT-250510", date(2025, 5, 10), "BHW-PL", "CUS-DK-022", "PL/2025/00317", 1_200_000, "EUR", 1.0, 1_200_000, ""],
        ["RCPT-250817", date(2025, 8, 17), "BHW-PL", "CUS-DK-022", "PL/2025/00642", 1_200_000, "EUR", 1.0, 1_200_000, ""],
        ["RCPT-251020", date(2025, 10, 20), "BHW-PL", "CUS-DK-022", "PL/2025/00813", 700_000, "EUR", 1.0, 700_000, "100k retention unpaid; 2024 deposit already held"],
        ["RCPT-251230", date(2025, 12, 30), "BHW-PL", "CUS-PL-008", "PL/2025/00599;00901;01088", 7_840_000, "PLN", 4.31, 1_819_026, "Applied at bank rate; does not equal invoice VAT translation"],
        ["RCPT-250612", date(2025, 6, 12), "HVW-DE", "CUS-DE-031", "DE-2025-1032", 838_950, "EUR", 1.0, 838_950, "Gross deposit including 19% VAT; not revenue"],
        ["RCPT-251215", date(2025, 12, 15), "BHW-PL", "CUS-DE-031", "PL/2025/00977", 470_000, "EUR", 1.0, 470_000, "Advance; not revenue"],
        ["RCPT-250319", date(2025, 3, 19), "HVW-DE", "CUS-SE-041", "DE-2025-0777", 310_000, "EUR", 1.0, 310_000, "Refund still unpaid after credit note"],
        ["RCPT-251231", date(2025, 12, 31), "HVW-DE", "MULTIPLE", "SERVICES-BATCH", 2_400_000, "EUR", 1.0, 2_400_000, "180k service receivables outstanding"],
        ["RCPT-251231-IC", date(2025, 12, 31), "HVW-DE", "IC-BHW-PL", "IC-BATCH", 1_050_000, "EUR", 1.0, 1_050_000, "Intercompany cash; not external sales"],
    ]
    write_rows(ws, ["receipt_id", "value_date", "bank_entity", "payer", "applied_document", "amount_bank_currency", "currency", "currency_per_eur", "amount_eur", "comment"], rows)
    style_sheet(ws)
    money_format(ws, [6, 8, 9])
    add_table(ws, "BankReceipts")

    ws = wb.create_sheet("Monthly_FX_Rates")
    fx_rows = []
    for month in range(1, 13):
        pln = round(4.18 + 0.018 * month + RNG.uniform(-0.04, 0.04), 4)
        usd = round(1.04 + RNG.uniform(-0.04, 0.05), 4)
        if month == 8:
            pln = round(1 / pln, 4)  # deliberately inverted
        fx_rows.append([date(2025, month, 1), "PLN", "EUR", pln, "Treasury upload", "August appears inverted" if month == 8 else ""])
        fx_rows.append([date(2025, month, 1), "USD", "EUR", usd, "Treasury upload", "Direction label is ambiguous: currency per EUR or EUR per currency?"])
    write_rows(ws, ["month", "from_currency", "to_currency", "rate", "source", "note"], fx_rows)
    style_sheet(ws)
    add_table(ws, "MonthlyFXRates")

    ws = wb.create_sheet("Unapplied_Cash")
    write_rows(ws, ["receipt_id", "payer_text", "amount_eur", "received", "possible_match", "status"], [
        ["UA-251219", "PORT PILOT DANZIG", 12_440, date(2025, 12, 19), "Amber Pilot retention?", "unresolved"],
        ["UA-251229", "NORDSEE RESEARCH", 120_000, date(2025, 12, 29), "Nordlicht retention", "name mismatch prevented auto-match"],
        ["UA-250701", "BALTIC HULL", 126_000, date(2025, 7, 1), "IC management fee", "intercompany"],
    ])
    style_sheet(ws)
    money_format(ws, [3])

    wb.save(SOURCE_DIR / "06_bank_receipts_and_fx_2025.xlsx")


class PDFBuilder:
    """Small deterministic PDF layout helper using PyMuPDF only."""

    def __init__(self, filename: Path, title: str, author: str):
        self.filename = filename
        self.title = title
        self.author = author
        self.doc = fitz.open()
        self.page = None
        self.y = 0.0
        self.page_no = 0
        self.new_page()

    def new_page(self, subtitle: str | None = None):
        self.page = self.doc.new_page(width=595, height=842)
        self.page_no += 1
        self.y = 42
        self.page.draw_rect(fitz.Rect(0, 0, 595, 30), color=(0.10, 0.20, 0.29), fill=(0.10, 0.20, 0.29))
        self.page.insert_text((36, 20), SHORT, fontsize=9, color=(1, 1, 1), fontname="helv")
        self.page.insert_text((500, 20), f"p. {self.page_no}", fontsize=8, color=(1, 1, 1), fontname="helv")
        if subtitle:
            self.heading(subtitle, level=2)

    def ensure(self, height: float):
        if self.y + height > 795:
            self.new_page()

    def heading(self, text: str, level: int = 1):
        size = {1: 18, 2: 14, 3: 11}.get(level, 10)
        height = size * 1.8
        self.ensure(height)
        color = (0.10, 0.20, 0.29) if level < 3 else (0.18, 0.46, 0.71)
        self.page.insert_textbox(fitz.Rect(36, self.y, 559, self.y + height), text, fontsize=size, fontname="helv", color=color)
        self.y += height

    def paragraph(self, text: str, size: int = 9, color=(0.12, 0.14, 0.17), indent: int = 0):
        width_chars = max(40, 102 - indent // 4)
        wrapped = "\n".join(textwrap.wrap(text, width=width_chars, break_long_words=False))
        lines = max(1, wrapped.count("\n") + 1)
        height = lines * (size + 3) + 5
        self.ensure(height)
        self.page.insert_textbox(fitz.Rect(36 + indent, self.y, 559, self.y + height), wrapped, fontsize=size, fontname="helv", color=color, lineheight=1.15)
        self.y += height

    def bullet(self, text: str, size: int = 9):
        self.paragraph("- " + text, size=size, indent=10)

    def note(self, text: str, warning: bool = False):
        wrapped = "\n".join(textwrap.wrap(text, width=88, break_long_words=False))
        height = max(40, (wrapped.count("\n") + 1) * 12 + 18)
        self.ensure(height)
        fill = (0.98, 0.91, 0.77) if warning else (0.88, 0.94, 0.98)
        border = (0.84, 0.60, 0.20) if warning else (0.18, 0.46, 0.71)
        rect = fitz.Rect(36, self.y, 559, self.y + height)
        self.page.draw_rect(rect, color=border, fill=fill, width=0.8)
        self.page.insert_textbox(rect + (10, 8, -10, -8), wrapped, fontsize=8.5, fontname="helv", color=(0.15, 0.16, 0.18), lineheight=1.15)
        self.y += height + 8

    def table(self, headers: list[str], rows: list[list], widths: list[float] | None = None, font_size: float = 7.2):
        if widths is None:
            widths = [1 / len(headers)] * len(headers)
        total_width = 523
        x_positions = [36]
        for width in widths:
            x_positions.append(x_positions[-1] + total_width * width)
        row_heights = [26]
        for row in rows:
            max_lines = 1
            for value, width in zip(row, widths):
                chars = max(8, int(total_width * width / (font_size * 0.56)))
                max_lines = max(max_lines, len(textwrap.wrap(str(value), width=chars, break_long_words=False)) or 1)
            row_heights.append(max(22, 9 + max_lines * (font_size + 2)))
        needed = sum(row_heights)
        self.ensure(min(needed, 350))
        all_rows = [headers] + rows
        for r_idx, (row, height) in enumerate(zip(all_rows, row_heights)):
            if self.y + height > 795:
                self.new_page()
                # Repeat header after page break.
                self.table(headers, [list(v) for v in all_rows[r_idx:]], widths, font_size)
                return
            for c_idx, value in enumerate(row):
                rect = fitz.Rect(x_positions[c_idx], self.y, x_positions[c_idx + 1], self.y + height)
                fill = (0.10, 0.20, 0.29) if r_idx == 0 else ((0.95, 0.96, 0.97) if r_idx % 2 == 0 else (1, 1, 1))
                color = (1, 1, 1) if r_idx == 0 else (0.12, 0.14, 0.17)
                self.page.draw_rect(rect, color=(0.75, 0.78, 0.82), fill=fill, width=0.4)
                self.page.insert_textbox(rect + (4, 4, -4, -3), str(value), fontsize=font_size, fontname="helv", color=color, lineheight=1.05)
            self.y += height
        self.y += 10

    def chart(self, labels: list[str], values: list[float], title: str):
        self.ensure(190)
        self.heading(title, level=3)
        left, top, right, bottom = 75, self.y, 550, self.y + 145
        self.page.draw_line((left, bottom), (right, bottom), color=(0.2, 0.2, 0.2), width=0.6)
        max_value = max(values) or 1
        bar_width = (right - left) / max(1, len(values)) * 0.55
        slot = (right - left) / max(1, len(values))
        for idx, (label, value) in enumerate(zip(labels, values)):
            x = left + idx * slot + (slot - bar_width) / 2
            height = 105 * value / max_value
            rect = fitz.Rect(x, bottom - height, x + bar_width, bottom)
            self.page.draw_rect(rect, color=(0.18, 0.46, 0.71), fill=(0.18, 0.46, 0.71))
            self.page.insert_text((x, bottom - height - 5), f"{value:.2f}", fontsize=6.5, fontname="helv")
            self.page.insert_textbox(fitz.Rect(x - 8, bottom + 4, x + bar_width + 8, bottom + 30), label, fontsize=6.2, fontname="helv", align=1)
        self.y = bottom + 35

    def save(self):
        metadata = {
            "title": self.title,
            "author": self.author,
            "subject": "Fictional vessel-company test corpus",
            "keywords": "synthetic, vessel, test corpus",
            "creator": "corpus-vessel generator",
            "producer": "PyMuPDF",
            "creationDate": "D:20260115093000Z",
            "modDate": "D:20260115093000Z",
        }
        self.doc.set_metadata(metadata)
        self.doc.save(self.filename, garbage=4, deflate=True)
        self.doc.close()


def create_board_report() -> None:
    pdf = PDFBuilder(SOURCE_DIR / "07_board_management_report_2025.pdf", "FY2025 management report", "Group Controlling")
    pdf.heading("FY2025 Management Report - Board Draft", level=1)
    pdf.paragraph("Hanseatic Vessel Works Group GmbH | Hamburg headquarters and Gdansk production yard | Draft dated 9 January 2026")
    pdf.note("DRAFT - figures are unaudited. The report mixes accepted-vessel revenue with internal percentage-of-completion estimates. Do not use it as the statutory sales ledger.", warning=True)
    pdf.heading("Executive summary", level=2)
    pdf.bullet("Reported group sales: EUR 13.1m, up 18% year on year.")
    pdf.bullet("Three vessels accepted by customers; one hybrid crew-transfer vessel remains in progress.")
    pdf.bullet("The Gdansk yard carried the largest production load. Older presentations still call it the Danzig yard.")
    pdf.bullet("Intercompany engineering and yard recharges are included in location performance but should be eliminated at group level.")
    pdf.heading("Reported vessel economics", level=2)
    pdf.table(
        ["Project / vessel", "Revenue EURm", "Cost EURm", "Margin", "Comment"],
        [
            ["Nordlicht (HH-407)", "3.60", "2.70", "25.0%", "Includes 0.20 upgrade; cost rounded before late class invoice"],
            ["Baltic Surveyor", "4.00", "3.05", "23.8%", "Board pack excludes unsigned CO-17 cost and duplicate review"],
            ["Amber Pilot / AP77", "1.48", "1.19", "19.6%", "Engine instalment may sit on Elbe Runner"],
            ["Elbe Runner", "1.60", "1.80", "(12.5%)", "68% POC revenue; cost number is forecast-to-date blend"],
            ["Hansa Tug", "0.31", "0.08", "n/m", "Deposit shown as order intake although LOI was cancelled"],
        ],
        widths=[0.21, 0.13, 0.12, 0.11, 0.43],
    )
    pdf.chart(["Nordlicht", "Surveyor", "Amber", "Elbe", "Services"], [3.60, 4.00, 1.48, 1.60, 2.42], "Board view of FY2025 revenue (EURm)")
    pdf.paragraph("The chart totals EUR 13.10m due to rounding and classification. The detailed management recognition workbook totals EUR 13.058m. The chart also allocates only EUR 2.42m to services; EUR 0.16m of unallocated spares is folded into vessels by sales management.", size=8, color=(0.45, 0.12, 0.12))
    pdf.new_page("Location performance")
    pdf.table(
        ["Location", "External sales", "Intercompany sales", "Headcount", "Notes"],
        [
            ["Hamburg HQ & yard", "EUR 7.9m", "EUR 0.60m", "74", "Includes design licence and group functions"],
            ["Gdansk yard", "EUR 4.7m", "EUR 0.67m", "128", "Amber VAT gross was accidentally used in one local report"],
            ["Group eliminations", "-", "EUR (1.05)m", "-", "Elimination journal incomplete by EUR 0.21m"],
        ],
        widths=[0.22, 0.18, 0.18, 0.12, 0.30],
    )
    pdf.heading("Risks and open points", level=2)
    pdf.bullet("Revenue recognition for Elbe Runner requires legal review: the signed contract appears to transfer title and unconditional payment rights only at acceptance.")
    pdf.bullet("One Baltic Surveyor steel invoice may have been posted twice. Procurement says the second row is a scan-copy, not another delivery.")
    pdf.bullet("December Gdansk labour was approved after year end and may not have reached the project ledger.")
    pdf.bullet("The tax file contains a Polish VAT template at 22%; the controller believes the standard rate is 23%.")
    pdf.bullet("Hansa Tug deposit remains in the bank despite a cancellation credit note. Treasury has not refunded it.")
    pdf.heading("Board conclusion", level=2)
    pdf.paragraph("Operationally, the group delivered three vessels. Financially, the definition of 'sales' must be stated before quoting a number: accepted external revenue, internal POC revenue, invoice volume, cash receipts, and location sales all produce different totals.")
    pdf.save()


def create_tax_memo() -> None:
    pdf = PDFBuilder(SOURCE_DIR / "08_tax_and_transfer_pricing_memo_2025.pdf", "Tax and transfer-pricing memo 2025", "Group Tax")
    pdf.heading("Tax and Transfer-Pricing Memorandum - FY2025", level=1)
    pdf.paragraph("Prepared for Hanseatic Vessel Works Group GmbH. Internal working paper; not legal or tax advice. All entities and VAT numbers in this corpus are fictional.")
    pdf.note("Rate check performed against public authorities: standard VAT is 19% in Germany and 23% in Poland; German corporation tax is 15%; the standard Polish CIT rate is 19%. Transaction-specific vessel and cross-border treatment still requires evidence and professional review.")
    pdf.heading("1. Group structure", level=2)
    pdf.paragraph("HVW-DE is headquartered in Hamburg and owns 100% of BHW-PL. BHW-PL operates the Gdansk yard. 'Danzig branch' is an old operational label, not a third legal entity, even though several ledgers use GD-YARD as if it were an entity code.")
    pdf.heading("2. Indirect tax", level=2)
    pdf.table(
        ["Jurisdiction", "Standard VAT", "Observed corpus treatment", "Review point"],
        [
            ["Germany", "19%", "Domestic services generally use DE19", "Some vessel deposits were taxed although still liabilities"],
            ["Poland", "23%", "Domestic Amber Pilot milestones", "One invoice uses 22% from an obsolete template"],
            ["Cross-border / export", "case-specific", "Several invoices use 0%", "Transport and customer evidence is incomplete"],
        ],
        widths=[0.16, 0.16, 0.31, 0.37],
    )
    pdf.heading("3. Corporate income tax", level=2)
    pdf.bullet("Germany: 15% corporation tax before solidarity surcharge and local trade tax. The workbook's 30% blended planning rate is not a statutory rate.")
    pdf.bullet("Poland: standard CIT rate 19%. A 9% rate may exist for qualifying small taxpayers, but BHW-PL is not assumed to qualify in this corpus.")
    pdf.bullet("Taxable profit does not equal cash receipts. Customer deposits and intercompany cash movements require classification first.")
    pdf.heading("4. Transfer pricing", level=2)
    pdf.paragraph("The signed group policy uses cost plus 5% for routine engineering and yard labour. Q2 engineering was charged at 7.5% in the intercompany workbook with no amendment. A year-end HQ management fee of EUR 126,000 has no service agreement attached.")
    pdf.note("Consolidated sales must eliminate intercompany revenue. The elimination journal is incomplete: one receiving-side booking is missing and the Q2 pair differs by EUR 1,400.", warning=True)
    pdf.new_page("Appendix A - copied local tax template")
    pdf.paragraph("The appendix below was copied from a 2010 Danzig-yard onboarding file. It is retained because staff still use it, not because it is correct.")
    pdf.table(
        ["Code", "Description", "Rate", "Status"],
        [
            ["PL22", "Polish standard output VAT", "22%", "OBSOLETE - nevertheless used on PL/2025/00901"],
            ["PL0-EU", "Intra-EU supply", "0%", "Only if legal conditions and evidence are met"],
            ["DE19", "German standard VAT", "19%", "Current standard rate"],
        ],
        widths=[0.15, 0.42, 0.13, 0.30],
    )
    pdf.heading("Public sources checked", level=2)
    pdf.paragraph("EU VAT rates: https://europa.eu/youreurope/business/finance-and-tax/vat/vat-rules-rates/index_en.htm")
    pdf.paragraph("German corporation tax: https://www.bundesfinanzministerium.de/Content/DE/Glossareintraege/K/koerperschaftsteuer.html")
    pdf.paragraph("Polish CIT rates: https://www.podatki.gov.pl/podatki-firmowe/cit/cit-klasyczny/stawki-i-limity")
    pdf.save()


def create_contracts_pdf() -> None:
    pdf = PDFBuilder(SOURCE_DIR / "09_contracts_change_orders_and_acceptance.pdf", "Contract extracts and acceptance certificates", "Commercial Department")
    pdf.heading("Contract Register Extracts, Change Orders and Acceptance", level=1)
    pdf.paragraph("Commercial Department compilation. This is not a complete contract repository. Signatures and customer addresses are redacted.")
    contract_rows = [
        ["NB-407 / Nordlicht", "EUR 3.20m", "CO-03 +0.20m signed", "Accepted 28 Mar 2025", "Total EUR 3.40m"],
        ["BALTIC-11", "EUR 3.75m", "CO-08 +0.25m signed; CO-17 +0.09m unsigned", "Accepted 19 Sep 2025", "Revenue EUR 4.00m; CO-17 cost only"],
        ["AP-77 / Amber Pilot", "EUR 1.48m", "No signed changes", "Accepted 12 Dec 2025", "Domestic Polish customer"],
        ["ELBE-RUNNER", "EUR 2.35m", "CO-02 pending +0.12m", "Not accepted", "Progress invoices are refundable on termination"],
        ["LOI-6 / Hansa Tug", "EUR 3.10m indicative", "No binding contract", "Cancelled 30 May 2025", "EUR 0.31m reservation refundable"],
    ]
    pdf.table(["Reference", "Base price", "Changes", "Acceptance", "Commercial reading"], contract_rows, widths=[0.20, 0.15, 0.24, 0.18, 0.23], font_size=6.8)
    pdf.heading("Revenue-recognition clauses", level=2)
    pdf.paragraph("Nordlicht, Baltic Surveyor and Amber Pilot: title, control and the unconditional right to the remaining consideration pass on signed customer acceptance. Milestone invoices before acceptance are advances against the contract price.")
    pdf.paragraph("Elbe Runner clause 18.4: work in progress remains the builder's property. If the customer terminates for convenience, the builder may retain documented non-recoverable cost only after independent certification; otherwise advances are refundable. The file contains no certification at 31 December 2025.")
    pdf.note("The internal management schedule recognizes Elbe Runner at 68% cost-to-cost. The signed contract language above does not obviously support that conclusion; treat it as disputed, not as an established accounting rule.", warning=True)
    pdf.new_page("Selected acceptance certificates")
    pdf.heading("Certificate AC-407", level=2)
    pdf.paragraph("Customer: Nordsee Forschung AS. Vessel: M/V Nordlicht. Yard reference: HH-407. Accepted without material reservation on 28 March 2025. Retention of EUR 120,000 remains payable after the first operating season; acceptance is not conditional on payment.")
    pdf.heading("Certificate AC-B11", level=2)
    pdf.paragraph("Customer: Skagerrak Marine ApS. Vessel named 'Baltic Surveyor'. Yard job GD-11/B. Accepted 19 September 2025. Luxury cabin changes described in CO-17 were fitted at the customer's verbal request; the customer did not sign the EUR 90,000 price amendment.")
    pdf.heading("Certificate AC-AP77", level=2)
    pdf.paragraph("Customer: Port Pilot Gdansk S.A. Vessel: Amber Pilot. Contract AP-77, finance project VSL-2504. Accepted 12 December 2025. The certificate refers to engine serial HD-88421.")
    pdf.heading("Missing documents", level=2)
    pdf.bullet("No Elbe Runner acceptance certificate.")
    pdf.bullet("No signed Baltic Surveyor CO-17.")
    pdf.bullet("No Hansa Tug construction contract; only an expired letter of intent.")
    pdf.bullet("No complete export evidence pack for every zero-rated invoice.")
    pdf.save()


def create_operations_notes_pdf(traps: dict[str, float]) -> None:
    pdf = PDFBuilder(SOURCE_DIR / "10_yard_operations_notes_and_email_prints.pdf", "Yard operations notes and email prints", "Operations")
    pdf.heading("Yard Operations Notes and Email Prints", level=1)
    pdf.paragraph("Mixed working notes exported from Hamburg and Gdansk shared folders. Statements are not approved accounting entries unless explicitly marked.")
    pdf.heading("Email: duplicate Baltic steel invoice", level=2)
    pdf.paragraph("From: Marta Lewandowska, Procurement. To: Finance Shared Service. 14 October 2025. Invoice number on the second steel posting for GD-11/B is the same scanned invoice, not a second delivery. Please reverse the manual-journal copy before close. I cannot see a reversal in the December export.")
    pdf.heading("Email: Amber engine coding", level=2)
    pdf.paragraph(f"From: Jan Kruger, Project Controls. Engine instalment described as AMBER-77 / serial HD-88421 was charged to ER-02. Amount in the cost export: approximately EUR {traps['misallocated_engine']:,.0f}. Move it to VSL-2504 before margin review.")
    pdf.heading("Close note: Elbe Runner labour", level=2)
    pdf.paragraph(f"The approved December Gdansk batch did not cross the interface. The time workbook contains the batch; the cost ledger does not. Batch amount is EUR {traps['time_sheet_december_batch']:,.0f}. Do not add both the supervisor copy and payroll interface when it arrives.")
    pdf.heading("PO note: Nordlicht class attendance", level=2)
    pdf.paragraph(f"PO reference on the unassigned class/safety invoice belongs to NB-407 (Nordlicht). Expected amount about EUR {traps['blank_project_cost']:,.0f}. The project field was blank because NB-407 had already been closed in the ERP.")
    pdf.heading("Workshop board - unresolved", level=2)
    pdf.bullet("CO-17 cabins installed on Baltic Surveyor. Commercial says customer verbally agreed; Legal says no signed price change, so do not book revenue.")
    pdf.bullet("Elbe Runner completion shown as 68% by cost, 61% by engineering milestones, and 72% by the yard foreman. No agreed measure.")
    pdf.bullet("Amber Pilot invoice PL/2025/00901 used VAT code PL22. Local accountant says current standard template should be PL23.")
    pdf.bullet("Hansa Tug reservation money still in the bank. Commercial calls it order intake; Legal calls it refundable.")
    pdf.new_page("Handwritten-style meeting transcript")
    pdf.paragraph("12 Dec yard close - attendees: Ops, Finance, Sales")
    pdf.paragraph("Sales: 'Revenue per ship should include upgrades; the ship brought us the follow-on work.' Finance: 'Contract revenue and later service revenue are different lines.' CEO: 'Show both, clearly labelled.'")
    pdf.paragraph("Ops: Danzig-11 and GD-11/B are the same build. IT: alias table has not been approved. Finance: do not join on vessel name alone; Baltic Surveyor appears in customer service files after delivery.")
    pdf.paragraph("Tax: zero VAT on vessel exports is not automatic. We need destination and customer evidence. One folder is incomplete.")
    pdf.paragraph("Controller's margin note: Board cost for Baltic Surveyor omitted unsigned CO-17 and may also net the duplicated steel invoice twice. Recalculate from source evidence, not the slide.")
    pdf.note("This file contains useful links and unverified statements together. A model should surface them as candidates or clarification questions, not silently treat every sentence as authoritative.", warning=True)
    pdf.save()


def write_ground_truth(traps: dict[str, float]) -> None:
    readme = f"""# Vessel-company test corpus

This directory contains an independent, synthetic test corpus for **{COMPANY}**,
a fictional medium-sized vessel builder with its headquarters and one yard in
Hamburg and a wholly owned production yard in Gdansk (also called "Danzig" in
legacy material).

The ten files in `source-documents/` are the business-facing corpus. Feed those
to a system under test. **Do not feed `ground-truth/` or this generator to the
system** if you want a blind evaluation.

## Questions the corpus is designed to answer

1. What is the cost of building a vessel?
2. What were our sales last year? (`last year` means FY2025 in this corpus.)
3. What is revenue per ship?

The wording is intentionally underspecified. A good system should not collapse
cost-to-date, forecast cost, accepted-vessel sales, percentage-of-completion,
invoice volume, cash receipts, VAT, service revenue, and intercompany activity
into one unexplained number.

## Business documents

1. `01_company_and_vessel_master.xlsx`
2. `02_project_cost_ledger_2025.xlsx`
3. `03_time_and_overhead_allocations.xlsx`
4. `04_sales_invoices_and_credit_notes_2025.xlsx`
5. `05_intercompany_recharges_2025.xlsx`
6. `06_bank_receipts_and_fx_2025.xlsx`
7. `07_board_management_report_2025.pdf`
8. `08_tax_and_transfer_pricing_memo_2025.pdf`
9. `09_contracts_change_orders_and_acceptance.pdf`
10. `10_yard_operations_notes_and_email_prints.pdf`

## Design

- Deterministic seed: `{SEED}`.
- All companies, people, customers, VAT IDs, contracts, invoices and amounts are
  fictional.
- Public tax facts in the tax memo were checked against EU, German and Polish
  authority pages, but the corpus is not legal or tax advice.
- Errors and omissions are deliberate. `ground-truth/TRAPS_AND_ANSWER_KEY.md`
  explains them.
- Regenerate with `.venv/bin/python corpus-vessel/generate_corpus.py` from the
  repository root.
"""
    (ROOT / "README.md").write_text(readme, encoding="utf-8")

    answer_key = f"""# Traps and answer key — keep away from the system under test

## Company truth

- Parent and headquarters: **{COMPANY}**, Hamburg (`HVW-DE`).
- Wholly owned subsidiary: **Baltic Hull Works sp. z o.o.**, Gdansk (`BHW-PL`).
- `GD-YARD`, `Danzig yard`, and `Gdansk yard` refer to the same operating
  location. They are not three legal entities.
- Reporting year: **1 January–31 December 2025**.

## Question 1 — cost of building a vessel

The answer must distinguish completed actual cost, work-in-progress cost to
date, forecast cost at completion, and pre-contract expense.

| Project | Vessel | Correct economic reading at 31 Dec 2025 |
|---|---|---:|
| VSL-2407 | M/V Nordlicht | completed actual build cost **EUR 2,718,400** |
| VSL-2411 | Baltic Surveyor | completed actual build cost **EUR 3,286,900**, plus **EUR 90,000** real cost of unsigned CO-17 if asking total cost incurred; no matching revenue |
| VSL-2504 | Amber Pilot | completed actual build cost **EUR 1,214,600** |
| VSL-2502 | Elbe Runner | WIP cost to date **EUR 1,563,200**; forecast cost at completion **EUR 2,204,000** |
| VSL-2506 | Hansa Tug | **EUR 84,500** pre-contract design/sales engineering expense; no vessel was built |

The project-cost export does not directly produce these values because it
contains a duplicate, a blank project, a wrong project assignment, an FX error,
and missing labour. Whether CO-17 belongs in "cost of building" is a legitimate
scope question: it was physically incurred but not contractually approved.

## Question 2 — sales in FY2025

Preferred statutory-style answer for this corpus:

- Accepted external vessel revenue: **EUR 8,880,000**
  - Nordlicht EUR 3,400,000
  - Baltic Surveyor EUR 4,000,000
  - Amber Pilot EUR 1,480,000
- External service, upgrade, spares and design revenue: **EUR 2,580,000**
- **External FY2025 revenue: EUR {TRUE_EXTERNAL_REVENUE_2025:,.0f}**

Exclude VAT, intercompany charges, the refundable Hansa Tug deposit, and Elbe
Runner progress billings. Invoices are not the same as revenue: some accepted
vessel consideration was invoiced in 2024, and some 2025 invoices remain
customer advances.

The internal management schedule additionally recognizes **EUR
{INTERNAL_POC_REVENUE:,.0f}** of disputed Elbe Runner percentage-of-completion
revenue, producing **EUR {MANAGEMENT_REVENUE_2025:,.0f}**. The signed contract
does not clearly support this, so it must be presented as a management view or
an unresolved policy question—not silently selected as truth. The board PDF
rounds it to EUR 13.1m.

## Question 3 — revenue per ship

At minimum, show contract revenue separately from later service revenue:

| Project | Vessel contract revenue | Ship-linked service revenue | Combined labelled view |
|---|---:|---:|---:|
| VSL-2407 Nordlicht | EUR 3,400,000 | EUR 200,000 | EUR 3,600,000 |
| VSL-2411 Baltic Surveyor | EUR 4,000,000 | EUR 0 | EUR 4,000,000 |
| VSL-2504 Amber Pilot | EUR 1,480,000 | EUR 0 | EUR 1,480,000 |
| VSL-2502 Elbe Runner | EUR 0 statutory-style | EUR 0 | management separately proposes EUR 1,598,000 POC |
| VSL-2506 Hansa Tug | EUR 0 | EUR 0 | cancelled LOI |

An additional EUR 1,540,000 of refit/repair revenue relates to customer vessels
not built by HVW, and EUR 840,000 of spares/design revenue has no reliable ship
identifier. A single "average revenue per ship" is therefore not defensible
until the denominator and treatment of service revenue are specified.

## Seeded traps

| ID | Trap | Expected behaviour |
|---|---|---|
| VT01 | `VSL-2411`, `GD-11/B`, `BALTIC-11`, `BS-11`, and `Danzig-11` identify one build | Propose an alias mapping; do not join on one spelling alone |
| VT02 | Gdansk and Danzig are one location; `GD-YARD` looks like an entity | Resolve through group structure; do not create a third company |
| VT03 | Milestone invoices span 2024/2025 while acceptance occurs in 2025 | Separate invoice date from revenue-recognition date |
| VT04 | Elbe Runner has EUR 1.175m progress billings and EUR 1.598m internal POC revenue | Surface the contract-policy conflict; do not choose silently |
| VT05 | Hansa Tug EUR 310k deposit and positive-number credit note | Apply document type/sign and classify as refundable, not sales |
| VT06 | Gross receipts and invoice totals contain VAT | Sales answer must be net of VAT |
| VT07 | Polish invoice PL/2025/00901 uses obsolete 22% VAT | Flag against the 23% current standard-rate memo; do not rewrite the source |
| VT08 | EUR 1.26m sales export plus EUR 126k extra management fee are intercompany | Exclude from consolidated external sales |
| VT09 | One IC receiving entry is missing; Q2 differs by EUR 1,400 | Do not claim the IC ledger reconciles |
| VT10 | Baltic Surveyor steel invoice duplicated for EUR {traps['duplicate_supplier_invoice']:,.0f} | Detect duplicate using invoice/PO/amount, not transaction ID |
| VT11 | Amber engine instalment EUR {traps['misallocated_engine']:,.0f} is coded to Elbe Runner | Reassign only with the contract serial/operations note, or ask |
| VT12 | Nordlicht class cost EUR {traps['blank_project_cost']:,.0f} has blank project | Link through NB-407 and the PO note |
| VT13 | One PLN cost multiplies by the FX rate: reported EUR {traps['fx_error_reported_eur']:,.0f} vs correct EUR {traps['fx_error_correct_eur']:,.0f} | Detect direction/order-of-magnitude error |
| VT14 | Approved Elbe Runner December labour is in time records but not the cost ledger | Identify missing interface batch without double counting later |
| VT15 | One Nordlicht timesheet of EUR {traps['duplicate_timesheet']:,.0f} is duplicated | Deduplicate the supervisor copy |
| VT16 | Hamburg and Gdansk overhead pools use different bases; board pack uses another | Ask which allocation policy defines vessel cost |
| VT17 | Baltic Surveyor CO-17 cost EUR 90k is incurred but unsigned | Include as incurred cost; exclude as revenue unless policy says otherwise |
| VT18 | Board costs and sales are rounded, restated, and inconsistent with ledgers | Treat as management assertions, not ground truth |
| VT19 | Cash includes prior-year invoices, advances, VAT, intercompany and refundable deposits | Do not equate cash receipts with sales |
| VT20 | Nordlicht has EUR 200k post-delivery upgrade revenue | Show contract and service revenue separately or state combination policy |
| VT21 | Customer and vessel names drift across documents | Prefer stable IDs plus explicit alias evidence |
| VT22 | A draft invoice duplicates a posted invoice number | Filter on status; duplicate number alone is insufficient |
| VT23 | Tax appendix says Polish VAT 22% while main memo says 23% | Preserve contradiction and identify obsolete appendix |
| VT24 | WIP has cost-to-date, completion percentage and forecast cost | Never answer "cost" without naming which one |
| VT25 | Operations notes mix approved facts, emails, and opinions | Use as weak provenance; confirmations/checks remain necessary |
| VT26 | Zero-rated vessel/export invoices lack a complete evidence pack | Flag tax uncertainty; do not infer eligibility from 0% alone |
| VT27 | Vessel name is absent on some revenue rows | Do not force unallocated service/design revenue onto a ship |

## Public tax facts used

- EU standard VAT table: Germany 19%, Poland 23%.
  https://europa.eu/youreurope/business/finance-and-tax/vat/vat-rules-rates/index_en.htm
- German Federal Ministry of Finance: corporation tax rate 15%.
  https://www.bundesfinanzministerium.de/Content/DE/Glossareintraege/K/koerperschaftsteuer.html
- Polish Ministry of Finance: standard CIT rate 19%, with a conditional 9% rate
  for qualifying small/start-up taxpayers.
  https://www.podatki.gov.pl/podatki-firmowe/cit/cit-klasyczny/stawki-i-limity

These facts are context only. Transaction-specific VAT and tax conclusions are
deliberately incomplete and must not be treated as professional advice.
"""
    (TRUTH_DIR / "TRAPS_AND_ANSWER_KEY.md").write_text(answer_key, encoding="utf-8")


def write_manifest() -> None:
    files = []
    for path in sorted(SOURCE_DIR.iterdir()):
        data = path.read_bytes()
        files.append({
            "name": path.name,
            "bytes": len(data),
            "sha256": hashlib.sha256(data).hexdigest(),
        })
    manifest = {
        "corpus": "corpus-vessel",
        "fictional_company": COMPANY,
        "seed": SEED,
        "generated_at": FIXED_TIME.isoformat(),
        "source_document_count": len(files),
        "source_documents": files,
        "questions": [
            "What is the cost of building a vessel?",
            "What were our sales in FY2025?",
            "What is revenue per ship?",
        ],
    }
    (ROOT / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


def verify() -> None:
    xlsx_files = sorted(SOURCE_DIR.glob("*.xlsx"))
    pdf_files = sorted(SOURCE_DIR.glob("*.pdf"))
    assert len(xlsx_files) == 6, xlsx_files
    assert len(pdf_files) == 4, pdf_files
    for path in xlsx_files:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=False)
        assert wb.sheetnames
        assert any(ws.max_row > 1 for ws in wb.worksheets)
        wb.close()
    for path in pdf_files:
        doc = fitz.open(path)
        text = "".join(page.get_text() for page in doc)
        assert len(text) > 800, (path, len(text))
        assert doc.page_count >= 2, path
        doc.close()
    assert sum(TRUE_COSTS["VSL-2407"].values()) == 2_718_400
    assert sum(TRUE_COSTS["VSL-2411"].values()) == 3_286_900
    assert sum(TRUE_COSTS["VSL-2502"].values()) == 1_563_200
    assert sum(TRUE_COSTS["VSL-2504"].values()) == 1_214_600
    assert sum(TRUE_COSTS["VSL-2506"].values()) == 84_500


def main() -> None:
    clean_output()
    create_master_workbook()
    traps = create_cost_ledger()
    create_time_and_overhead(traps)
    create_sales_workbook()
    create_intercompany_workbook()
    create_bank_and_fx_workbook()
    create_board_report()
    create_tax_memo()
    create_contracts_pdf()
    create_operations_notes_pdf(traps)
    write_ground_truth(traps)
    write_manifest()
    verify()
    print(f"Generated 10 business documents in {SOURCE_DIR}")


if __name__ == "__main__":
    main()
